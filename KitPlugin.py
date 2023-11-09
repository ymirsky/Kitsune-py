import csv
import math
import os

import openpyxl
import pandas as pd
from matplotlib import pyplot as plt
from optuna_dashboard import run_server

from Kitsune import Kitsune
from KitNET.KitNET import KitNET
import netStat as ns
import shap
import numpy as np
import pickle
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import sklearn
import optuna
from scipy.stats import norm
import random
from scapy.all import PcapReader, PcapWriter, wrpcap, rdpcap, IP, TCP, UDP

# Class that provides a callable interface for Kitsune components.
# Note that this approach nullifies the "incremental" aspect of Kitsune and significantly slows it down.
class KitPlugin:
    # Function used by SHAP as callback to test instances of features
    def kitsune_model(self, input_data):
        prediction = self.K.feed_batch(input_data)
        return prediction

    def kitnet_model(self, input_data):
        prediction = self.KitTest.process_batch(input_data)
        return prediction

    # Builds a Kitsune instance. Does not train KitNET yet.
    def __init__(self, input_path=None, packet_limit=None, num_autenc=None, FMgrace=None, ADgrace=None, learning_rate=0.1, hidden_ratio=0.75):
        # This code will be removed when batch running Kitsune has been finalized
        if input_path != None and num_autenc != None:
            self.features_list = None
            self.explainer = None
            self.shap_values = None
            self.K = Kitsune(input_path, packet_limit, num_autenc, FMgrace, ADgrace, learning_rate, hidden_ratio)
            self.metadata = {
                "filename" : input_path,
                "packet_limit" : packet_limit,
                "num_autenc" : num_autenc,
                "FMgrace": FMgrace,
                "ADgrace" : ADgrace,
                "timestamp" : datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
        self.testFeatures = None
        maxHost = 100000000000
        maxSess = 100000000000
        self.nstat = ns.netStat(np.nan, maxHost, maxSess)

    # Calls Kitsune's get_feature_list function to build the list of features
    def feature_builder(self, csv=False):
        print("Building features")
        # Dummy-running Kitsune to get a list of features
        self.features_list = self.K.get_feature_list(csv)
        return self.features_list

    # Loads Kitsune's feature list from a pickle file
    def feature_loader(self, newpickle=None):
        print("Loading features from file")
        path = 'pickles/featureList.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'rb') as f:
            features_list = pickle.load(f)
        self.features_list = features_list

    # Writes Kitsune's feature list to a pickle file
    def feature_pickle(self, newpickle=None):
        print("Writing features to file")
        path = 'pickles/featureList.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'wb') as f:
            pickle.dump(self.features_list, f)

    # Trains KitNET, using the specified index range of this class' feature list
    def kit_trainer(self, min_index, max_index):
        print("Training")
        self.K.feed_batch(self.features_list[min_index:max_index])
        print("Training finished")

    # Trains KitNET, using a supplied feature list
    def kit_trainer_supplied_features(self, features_list):
        print("Training")
        self.K.feed_batch(features_list)
        print("Training finished")

    # Runs KitNET, using specified index range of this class' feature list
    def kit_runner(self, min_index, max_index, normalize=False):
        print("Running")
        print(len(self.features_list[min_index:max_index]))
        return self.K.feed_batch(self.features_list[min_index:max_index])

    # Calculates KitNET's SHAP-values for the specified indexes
    def shap_values_builder(self, min_train, max_train, min_test, max_test):
        self.metadata['min_train'] = min_train
        self.metadata['max_train'] = max_train
        self.metadata['min_test'] = min_test
        self.metadata['max_test'] = max_test
        print("Building SHAP explainer")
        self.explainer = shap.Explainer(self.kitsune_model, np.array(self.features_list[min_train:max_train]))
        print("Calculating SHAP values")
        if self.testFeatures != None:
            self.shap_values = self.explainer.shap_values(np.array(self.testFeatures[min_test:max_test]))
        else:
            self.shap_values = self.explainer.shap_values(np.array(self.features_list[min_test:max_test]))
        return self.shap_values

    # Writes the SHAP-values to a pickle-file
    def shap_values_pickle(self, newpickle=None):
        path = 'pickles/shap_values.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'wb') as f:
            pickle.dump(self.shap_values, f)

    # Gets the SHAP-values from a pickle-file
    def shap_values_loader(self, newpickle=None):
        path = 'pickles/shap_values.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'rb') as f:
            self.shap_values = pickle.load(f)
        return self.shap_values

    # Calculates summary statistics of SHAP-values
    def shap_stats_summary_builder(self, min_index, max_index, plot_type="dot"):
        return shap.summary_plot(self.shap_values, np.array(self.features_list[min_index:max_index]), plot_type=plot_type)

    # Creates an Excel-file containing summary statistics for each feature
    def shap_stats_excel_export(self, path=None):
        self.workbook = openpyxl.load_workbook('input_data/template_statistics_file.xlsx')
        self.create_sheet("malicious_shap")
        excel_file = "summary_statistics_test.xlsx"
        if path != None:
            excel_file = path
        self.workbook.save(excel_file)
        print('done')

    # Calculates the three best and worst values for all statistics
    def get_high_low_indices(self):
        shap_transposed = self.shap_values.T
        # List of statistics functions
        stat_functions = {
            'mean': np.mean,
            'median': np.median,
            'std_dev': np.std,
            'variance': np.var,
            'minimum': np.min,
            'maximum': np.max,
            'total_sum': np.sum
        }

        # Dictionary to store results
        result_dict = {}

        # Loop over statistics
        for stat_name, stat_func in stat_functions.items():
            # Calculate the statistic for each list
            stat_values = stat_func(shap_transposed, axis=1)

            # Calculate the indices of the highest and lowest values
            sorted_indices = np.argsort(stat_values)
            highest_indices = sorted_indices[-3:]
            lowest_indices = sorted_indices[:3]
            # Store the indices in the result dictionary
            result_dict[stat_name] = {
                'highest_indices': highest_indices,
                'lowest_indices': lowest_indices
            }
        return result_dict

    def create_histogram(self, day, featuremean, featuremedian, sheet_title):
        # Extract keys and values from dictionaries
        keys = list(featuremean.keys())
        values_mean = list(featuremean.values())
        values_median = list(featuremedian.values())

        # Set up the figure and axis
        fig, ax = plt.subplots()

        # Set bar width
        bar_width = 0.35

        # Set the bar positions
        index = np.arange(len(keys))

        # Plot bars for featuremean
        bar1 = ax.bar(index, values_mean, bar_width, color='blue', label='Feature Mean')

        # Plot bars for featuremedian
        bar2 = ax.bar(index + bar_width, values_median, bar_width, color='orange', label='Feature Median')

        # Set labels and title
        ax.set_xlabel('Feature')
        plt.xticks(rotation=45)
        plt.tight_layout()
        ax.set_ylabel('SHAP-value')
        ax.set_title(f"{day}: {sheet_title}")
        ax.set_xticks(index + bar_width / 2)
        ax.set_xticklabels(keys)

        # Add legend
        ax.legend()
        plt.savefig(f'output_data/attack_types/{day}_{sheet_title}')
        # Show the plot
        plt.show()

    # Creates an Excel sheet with relevant statistics
    def create_sheet(self, day, sheet_title):
        sheet = self.workbook.copy_worksheet(self.workbook.active)
        sheet.title = sheet_title
        headers = ['Mean', 'Median', 'Standard Deviation', 'Variance', 'Minimum', 'Maximum', 'Sum', 'Metadata']
        header_row = headers
        lambdameans = {
            '5':[],
            '3':[],
            '1':[],
            '0.1':[],
            '0.01':[]
        }
        lambdamean = {
            '5': None,
            '3': None,
            '1': None,
            '0.1': None,
            '0.01': None
        }
        lambdamedians = {
            '5':[],
            '3':[],
            '1':[],
            '0.1':[],
            '0.01':[]
        }
        lambdamedian = {
            '5': None,
            '3': None,
            '1': None,
            '0.1': None,
            '0.01': None
        }
        featuremeans = {
            'weight':[],
            'mean':[],
            'standard deviation':[],
            'radius':[],
            'magnitude':[],
            'covariance':[],
            'pearson correlation coefficient':[]
        }
        featuremean = {
            'weight': None,
            'mean': None,
            'standard deviation': None,
            'radius': None,
            'magnitude': None,
            'covariance': None,
            'pearson correlation coefficient': None
        }
        featuremedians = {
            'weight':[],
            'mean':[],
            'standard deviation':[],
            'radius':[],
            'magnitude':[],
            'covariance':[],
            'pearson correlation coefficient':[]
        }
        featuremedian = {
            'weight': None,
            'mean': None,
            'standard deviation': None,
            'radius': None,
            'magnitude': None,
            'covariance': None,
            'pearson correlation coefficient': None
        }
        for col, value in enumerate(header_row):
            cell = sheet.cell(row=1, column=6 + col)
            cell.value = value
        for idx, num_list in enumerate(self.shap_values.T):
            mean = np.mean(num_list)
            median = np.median(num_list)
            std_dev = np.std(num_list)
            variance = np.var(num_list)
            minimum = np.min(num_list)
            maximum = np.max(num_list)
            total_sum = np.sum(num_list)
            if idx+1 <= 20:
                lambdameans['5'].append(mean)
                lambdamedians['5'].append(median)
            if idx+1 > 20 and idx+1 <= 40:
                lambdameans['3'].append(mean)
                lambdamedians['3'].append(median)
            if idx+1 > 40 and idx+1 <= 60:
                lambdameans['1'].append(mean)
                lambdamedians['1'].append(median)
            if idx+1 > 60 and idx+1 <= 80:
                lambdameans['0.1'].append(mean)
                lambdamedians['0.1'].append(median)
            if idx+1 > 80 and idx+1 <= 100:
                lambdameans['0.01'].append(mean)
                lambdamedians['0.01'].append(median)

            if idx+1 in [1, 4, 11, 14, 21, 24, 31, 34, 41, 44, 51, 54, 61, 64, 71, 74, 81, 84, 91, 94]:
                featuremeans['weight'].append(mean)
                featuremedians['weight'].append(median)
            if idx+1 in [2, 5, 12, 15, 22, 25, 32, 35, 42, 45, 52, 55, 62, 65, 72, 75, 82, 85, 92, 95]:
                featuremeans['mean'].append(mean)
                featuremedians['mean'].append(median)
            if idx + 1 in [3, 6, 13, 16, 23, 26, 33, 36, 43, 46, 53, 56, 63, 66, 73, 76, 83, 86, 93, 96]:
                featuremeans['standard deviation'].append(mean)
                featuremedians['standard deviation'].append(median)
            if idx + 1 in [7, 17, 27, 37, 47, 57, 67, 77, 87, 97]:
                featuremeans['radius'].append(mean)
                featuremedians['radius'].append(median)
            if idx + 1 in [8, 18, 28, 38, 48, 58, 68, 78, 88, 98]:
                featuremeans['magnitude'].append(mean)
                featuremedians['magnitude'].append(median)
            if idx + 1 in [9, 19, 29, 39, 49, 59, 69, 79, 89, 99]:
                featuremeans['covariance'].append(mean)
                featuremedians['covariance'].append(median)
            if idx + 1 in [10, 20, 30, 40, 50, 60, 70, 80, 99, 100]:
                featuremeans['pearson correlation coefficient'].append(mean)
                featuremedians['pearson correlation coefficient'].append(median)
            row_data = [mean, median, std_dev, variance, minimum, maximum, total_sum]

            for col, value in enumerate(row_data):
                cell = sheet.cell(row=idx + 2, column=6 + col)
                cell.value = value

        row = idx + 2
        row += 1
        cell = sheet.cell(row=row, column=1)
        cell.value = "Grouped by lambda"
        cell = sheet.cell(row=row, column=2)
        cell.value = "mean"
        cell = sheet.cell(row=row, column=3)
        cell.value = "median"
        row += 1

        for key in lambdameans:
            cell = sheet.cell(row=row, column = 1)
            cell.value = key
            cell = sheet.cell(row=row, column = 2)
            cell.value = np.mean(np.array(lambdameans[key]))
            lambdamean[key] = np.mean(np.array(lambdameans[key]))
            cell = sheet.cell(row=row, column=3)
            cell.value = np.median(np.array(lambdamedians[key]))
            lambdamedian[key] = np.mean(np.array(lambdamedians[key]))
            row += 1
        row += 1

        cell = sheet.cell(row=row, column=1)
        cell.value = "Grouped by feature"
        cell = sheet.cell(row=row, column=2)
        cell.value = "mean"
        cell = sheet.cell(row=row, column=3)
        cell.value = "median"
        row += 1
        for key in featuremeans:
            cell = sheet.cell(row=row, column=1)
            cell.value = key
            cell = sheet.cell(row=row, column=2)
            cell.value = np.mean(np.array(featuremeans[key]))
            featuremean[key] = np.mean(np.array(featuremeans[key]))
            cell = sheet.cell(row=row, column=3)
            cell.value = np.median(np.array(featuremedians[key]))
            featuremedian[key] = np.median(np.array(featuremedians[key]))
            row += 1
        self.create_histogram(day, featuremean, featuremedian, sheet_title+" grouped by feature name")
        self.create_histogram(day, lambdamean, lambdamedian, sheet_title + " grouped by lambda value")
        row += 1

        color_indices = self.get_high_low_indices()
        stat_columns = {
            'mean': "F",
            'median': "G",
            'std_dev': "H",
            'variance': "I",
            'minimum': "J",
            'maximum': "K",
            'total_sum': "L"
        }
        for stat in color_indices:
            for index in color_indices[stat]["highest_indices"]:
                cell_index = stat_columns[stat] + str(index + 2)
                if stat == "std_dev" or stat == "variance":
                    # Make largest three standard deviation and variance values blue
                    sheet[cell_index].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif stat == "minimum":
                    # Make largest three cells minimum red
                    sheet[cell_index].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                else:
                    # In all other cases, make largest three cells green
                    sheet[cell_index].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for index in color_indices[stat]["lowest_indices"]:
                cell_index = stat_columns[stat] + str(index + 2)
                if stat == "minimum":
                    # Make largest three cells minimum red
                    sheet[cell_index].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                else:
                    # In all other cases, make smallest three cells red
                    sheet[cell_index].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        # Fill in metadata
        start_row = 2
        start_column_keys = 'M'
        start_column_values = 'N'

        # Loop over the dictionary and write capitalized keys and values to cells
        if hasattr(self, "metadata"):
            for idx, (key, value) in enumerate(self.metadata.items()):
                capitalized_key = key[0].upper() + key[1:]
                key_cell = f"{start_column_keys}{start_row + idx}"
                value_cell = f"{start_column_values}{start_row + idx}"
                sheet[key_cell] = capitalized_key
                sheet[value_cell] = value
        return sheet

    # Runs a series of Kitsune models and calculates statistics for each run.
    def run_series_stats(self, inputs):
        self.workbook = openpyxl.load_workbook('input_data/template_statistics_file.xlsx')
        # Loop over the different Kitsune configs we are going to make
        for session in inputs:
            self.features_list = None
            self.explainer = None
            self.shap_values = None
            self.K = Kitsune(inputs[session]["input_path"], inputs[session]["packet_limit"], inputs[session]["maxAE"], inputs[session]["FMgrace"], inputs[session]["ADgrace"])
            self.metadata = {
                "filename": inputs[session]["input_path"],
                "packet_limit": inputs[session]["packet_limit"],
                "maxAE": inputs[session]["maxAE"],
                "FMgrace": inputs[session]["FMgrace"],
                "ADgrace": inputs[session]["ADgrace"],
                "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
            self.feature_builder()
            self.kit_trainer(inputs[session]["training_min"], inputs[session]["training_max"])
            if inputs[session]["input_path"] != inputs[session]["input_path_test"]:
                self.testKit = Kitsune(inputs[session]["input_path_test"], inputs[session]["packet_limit"], inputs[session]["maxAE"], inputs[session]["FMgrace"], inputs[session]["ADgrace"])
                self.testFeatures = self.testKit.get_feature_list()
            self.shap_values_builder(inputs[session]["training_min"], inputs[session]["training_max"], inputs[session]["testing_min"], inputs[session]["testing_max"])
            self.create_sheet(session)
        excel_file = "summary_statistics_" + datetime.now().strftime('%d-%m-%Y_%H-%M') + ".xlsx"
        self.workbook.save(excel_file)

    # Runs a hyperparameter optimization on the supplied dataset, constrained by number of runs and packet limit
    def hyper_opt(self, input_path, runs, packet_limit, load=False):
        if load:
            self.feature_loader()
        else:
            self.K = Kitsune(input_path, packet_limit * 1.3, 10, 5000, 50000, 0.1, 0.75)
            self.feature_builder()
            self.feature_pickle()

        def objective(trial):
            numAE = trial.suggest_int('numAE', 1, 10)
            learning_rate = trial.suggest_float('learning_rate', 0.01, 0.5)
            hidden_ratio = trial.suggest_float('hidden_ratio', 0.5, 0.8)

            self.K = Kitsune(input_path, packet_limit*1.3, numAE, int(0.1*packet_limit), int(0.6*packet_limit), learning_rate, hidden_ratio)
            # Load the feature list beforehand to save time
            self.feature_loader()
            print('training on '+str(int(0.7*packet_limit))+' packets')
            self.kit_trainer(0, int(0.7*packet_limit))

            y_test = np.zeros((int(0.2*packet_limit), 1))
            y_pred = self.kit_runner(int(0.7*packet_limit), int(0.9*packet_limit))

            # Do small test run with benign sample to find normalization
            print("Calculating normalization sample")
            #benignSample = np.log(self.kit_runner(int(0.5*packet_limit), int(0.6*packet_limit)))
            #logProbs = norm.logsf(np.log(y_pred), np.mean(benignSample), np.std(benignSample))
            print('predictions')
            print(y_pred)
            #print('normalization sample')
            #print(benignSample)
            #print('logProbs')
            #print(logProbs)
            error = sklearn.metrics.mean_squared_error(y_test, y_pred)

            print('error')
            print(error)
            return error

        study = optuna.create_study()
        study.optimize(objective, n_trials=runs)

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write header row
        header = ["Trial Number", "numAE", "learning_rate", "hidden_ratio"]
        ws.append(header)

        # Write trial information
        best_value = float("inf")
        best_row_idx = None  # Track the index of the best row
        for idx, trial in enumerate(study.trials, start=2):  # Start from row 2 to leave room for the header
            trial_params = trial.params
            trial_row = [trial.number, trial_params["numAE"], trial_params["learning_rate"], trial_params["hidden_ratio"], trial.value]
            ws.append(trial_row)

            if trial.value < best_value:
                best_value = trial.value
                best_row_idx = idx

        # Set fill color for the best value row
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        if best_row_idx is not None:
            for cell in ws[best_row_idx]:
                cell.fill = green_fill

        # Save the workbook to a file
        excel_file_path = "output_data/hyperparameter_optimization_results_" + datetime.now().strftime('%d-%m-%Y_%H-%M') + ".xlsx"
        wb.save(excel_file_path)

        print("Results exported to", excel_file_path)
        return study.best_trial

    # Calculates an EER-score for a list of RMSEs
    def calc_eer(self, RMSEs, labels):
        fpr, tpr, threshold = sklearn.metrics.roc_curve(labels, RMSEs, pos_label=1)
        fnr = 1-tpr
        #eer_threshold = threshold[np.nanargmin(np.absolute((fnr-fpr)))]
        EER = fpr[np.nanargmin(np.absolute((fnr-fpr)))]
        return EER

    # Calculates an AUC-score for a list of RMSEs and a list of expected values
    def calc_auc(self, RMSEs, labels):
        auc_score = sklearn.metrics.roc_auc_score(labels, RMSEs)
        return auc_score

    # Calculates an EER-score for a list of RMSEs and a list of expected values
    def calc_auc_eer(self, RMSEs, labels):
        return (self.calc_auc(RMSEs, labels), self.calc_eer(RMSEs, labels))

    # DEPRECATED Takes a random sample from a .pcap file, limited by the supplied sample size
    #def random_sample_pcap(self, input_path, output_path, sample_size):
    #    # Initialize the sampled_packets list and a counter
    #    sampled_packets = []
    #    counter = 0

        # Open the PCAP file for reading
    #    with PcapReader(input_path) as pcap_reader:
    #        for packet in pcap_reader:
    #            counter += 1
    #            if counter % 10000 == 0:
    #                print(counter)
    #            if len(sampled_packets) < sample_size:
    #                sampled_packets.append(packet)
    #            else:
    #                # Randomly decide whether to add the new packet or not
    #                probability = sample_size / counter
    #                if random.random() < probability:
    #                    random_index = random.randint(0, sample_size - 1)
    #                    sampled_packets[random_index] = packet

        # Write the sampled packets to a new PCAP file while preserving the order
    #    wrpcap(output_path, sampled_packets)

    #    print(f"Sampled {sample_size} packets and saved to {output_path}")

    # DEPRECATED Takes the first n percentage out of every 1000 packets, does the same for the next 1000 packets
    #def interval_sample_pcap(self, input_path, output_path, percentage):
    #    # Initialize the sampled_packets list and a counter
    #    sampled_packets = []
    #    counter = 0

    #    # Open the PCAP file for reading
    #    with PcapReader(input_path) as pcap_reader:
    #        for packet in pcap_reader:
    #            counter += 1
    #            if counter % 10000 == 0:
    #                print(counter)

    #            if counter % 1000 <= (1000*(percentage/100)):  # Sample the first 100 out of every 1000 packets
    #                sampled_packets.append(packet)

        # Write the sampled packets to a new PCAP file while preserving the order
    #    wrpcap(output_path, sampled_packets)

    #    print(f"Sampled the first 100 packets out of every 1000 and saved to {output_path}")

    # DEPRECATED Extracts the conversations from a pcap-file
    #def extract_conversations(self, input_path):
    #    print('Reading pcap-file')
    #    conversations = []
    #    current_conversation = []
    #    counter = 0

    #    with PcapReader(input_path) as pcap_reader:
    #        for packet in pcap_reader:
    #            counter += 1
    #            if counter % 10000 == 0:
    #                print(f"{counter} packets processed")

    #            if IP in packet:
    #                if TCP in packet:
    #                    conversation_key = (packet[IP].src, packet[IP].dst, packet[TCP].sport, packet[TCP].dport)
    #                elif UDP in packet:
    #                    conversation_key = (packet[IP].src, packet[IP].dst, packet[UDP].sport, packet[UDP].dport)
    #                else:
    #                    continue

    #                if conversation_key not in current_conversation:
    #                    current_conversation.append(conversation_key)
    #                    conversations.append([])

    #                conversations[current_conversation.index(conversation_key)].append(packet)

    #    self.conversations_list = conversations
    #    return conversations

    # DEPRECATED Writes a list of conversations to a pcap-file
    #def create_pcap_from_conversations(self, conversations, output_path):
    #    print('Writing packets to pcap-file')
    #    packets_to_write = []

    #    for conversation in conversations:
    #        packets_to_write.extend(conversation)

    #    with PcapWriter(output_path) as pcap_writer:
    #        pcap_writer.write(packets_to_write)

    # Sample a percentage of conversations (not of packets)
    # def sample_percentage_conversations(self, percentage, input_path, output_path=None):
    #    conversation_list = self.extract_conversations(input_path)
    # print(f'Sampling {percentage} percent of conversations')
    #    sampled_conversations = random.sample(conversation_list, int(0.01 * percentage * len(conversation_list)))

    #    if output_path is not None:
    #        self.create_pcap_from_conversations(sampled_conversations, output_path)

    #    self.conversations_list = sampled_conversations
    #    return sampled_conversations

    # DEPRECATED Trains Kitsune on a list of conversations
    # def train_Kitsune_on_conversations(self, conversation_list):
    #    self.K = Kitsune("input_data/empty.pcap", np.Inf, 6, math.floor(len(conversation_list)*0.1), math.floor(len(conversation_list)*0.9))
    #    for conversation in conversation_list:
    #        self.K.feed_batch(conversation)

    # DEPRECATED Runs Kitsune on a list of conversations and returns a list of anomaly-scores per conversation
    #def run_Kitsune_on_conversations(self, conversation_list, threshold):
    #    result_list = []
    #    malicious = 0
    #    for conversation in conversation_list:
    #        result = self.K.feed_batch(conversation)
    #        # Normalize result if maximum is a positive
    #        if max(result) >= 1.0:
    #            result = [float(i) / max(result) for i in result]
    #        # If one of the results is higher than the threshold, then mark as malicious
    #        if max(result) > threshold:
    #            malicious = 1
    #        # Add a tuple of conversation and malicious/benign
    #        result_list.append((conversation, malicious))
    #    return result_list

    # Loads conversations list from a pickle file
    def conversations_loader(self, newpickle=None):
        print("Loading conversations from file")
        path = 'pickles/conversationsList.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'rb') as f:
            conversations_list = pickle.load(f)
        self.conversations_list = conversations_list
        return conversations_list

    # Writes conversation list to a pickle file
    def conversation_pickle(self, newpickle=None):
        print("Writing conversations to file")
        path = 'pickles/conversationsList.pkl'
        if newpickle != None:
            path = newpickle
        with open(path, 'wb') as f:
            pickle.dump(self.conversations_list, f)

    # DEPRECATED Verifies a batch of conversations to be benign or malicious
    #def verify_test_results(self, conv_list, threshold):
    #    result_list = []
    #    for conv in conv_list:
    #        # If one of the results is higher than the threshold, then mark as malicious
    #        malicious = 0
    #        if max(conv[1]) > threshold:
    #            malicious = 1
    #        result_list.append((conv[0], malicious))
    #    return result_list

    def load_pcap_to_features(self, input_path):
        print('Running dummy instance of Kitsune')
        dummyKit = Kitsune(input_path, np.Inf, 6, 10, 15)
        self.features_list = dummyKit.get_feature_list()
        return self.features_list

    def read_label_file(self, csvpath):
        with open(csvpath, newline='') as csvfile:
            returnList = []
            labelreader = csv.reader(csvfile, delimiter=' ')
            for row in labelreader:
                row = row[0].strip('][').split(',')
                returnList.append(row)
            return returnList

    def sample_packets_by_conversation(self, tsvpath, outpath, labels):
        # We open the output writer to write to a new TSV file
        with open(outpath, 'w') as op:
            wr = csv.writer(op)
            # We open the reader to get the packets from the original TSV file
            with open(tsvpath) as fd:
                rd = csv.reader(fd, delimiter="\t", quotechar='"')
                pkt_iter = -1
                for row in rd:
                    if pkt_iter % 100000 == 0:
                        print(pkt_iter)
                    if pkt_iter == -1:
                        pkt_iter += 1
                        continue
                    # Labels is the list of conversations, that has previously been sampled to 10 percent of conversations
                    for label in labels:
                        if label[0] == 'Src':
                            continue
                        if (row[4] == label[0] and row[6] == label[1] and row[5] == label[2] and row[7] == label[3]) or (row[4] == label[2] and row[6] == label[3] and row[5] == label[0] and row[7] == label[1]):
                            label_iter = label[5]
                            label_val = label[4]
                            row.append(str(pkt_iter))
                            row.append(str(label_iter))
                            row.append(str(label_val))
                            wr.writerow(row)
                            break
                    pkt_iter += 1
            op.close()

    def map_packets_to_features(self, packet_path, feature_path, sampled_feature_path):
        # Step 1: Read the packet TSV file and create a set of packet indices
        subset_indices = set()
        row_index = 0
        with open(packet_path, 'r', newline='') as packet_file:
            csvreader = csv.reader(packet_file)
            for row in csvreader:
                if row:
                    #packet_index = int(row[19])  # Assuming index is in the 20th column
                    packet_index = int(row[22])  # Assuming index is in the 23rd column
                    subset_indices.add(packet_index)
                row_index += 1
        # Step 2: Read the required statistics from the large feature CSV file
        # and write them to the output CSV file
        with open(feature_path, 'r', newline='') as feature_file, open(sampled_feature_path, 'w', newline='') as output_file:
            csvreader = csv.reader(feature_file)
            csvwriter = csv.writer(output_file)
            print('thursday')

            for row_num, row in enumerate(csvreader, start=1):
                packet_index = row_num  # Index is the row number
                # Check if the packet index is in the list of subset indices
                if packet_index in subset_indices:
                    # Write the row to the output CSV file
                    csvwriter.writerow(row)

    # Runs a hyperparameter optimization on the supplied dataset, constrained by number of runs and packet limit
    # This version uses KitNET directly instead of running Kitsune as a whole
    def hyper_opt_KitNET(self, feature_path, training_cutoff):
        def objective(trial):
            numAE = trial.suggest_int('numAE', 0, 200)
            learning_rate = trial.suggest_float('learning_rate', 0, 0.5)
            hidden_ratio = trial.suggest_float('hidden_ratio', 0, 1)
            FMgrace = trial.suggest_int('FMgrace', 0, 500000)

            kit = KitNET(100, max_autoencoder_size=numAE, FM_grace_period=FMgrace, AD_grace_period=math.floor(training_cutoff*0.9), learning_rate=learning_rate, hidden_ratio=hidden_ratio)
            # Load the feature list beforehand to save time
            counter = 0

            with open(feature_path) as fp:
                rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
                train_err = []
                for packet in rd_ft:
                    if packet:
                        packet = packet[0].split(',')
                        packet = [float(element) for element in packet]
                        packet = np.array(packet)
                        if counter % 10000 == 0:
                            print("training: "+str(counter))
                        if counter <= training_cutoff:
                            train_err.append(kit.train(packet))
                        else:
                            break
                        counter += 1
                fp.close()

            y_pred = []
            path = 'pickles/validateFeatureList.pkl'
            counterValidate = 0
            with open(path, 'rb') as f:
                validateList = pickle.load(f)
            for packet in validateList:
                score = kit.execute(packet)
                if counterValidate % 100000:
                    print("testing: "+str(counterValidate))
                y_pred.append(score)

            trial.set_user_attr("training_error", np.mean(train_err))
            trial.set_user_attr("train_median", np.median(train_err))
            trial.set_user_attr("train_25_percentile", np.percentile(train_err, 25))
            trial.set_user_attr("train_75_percentile", np.percentile(train_err, 75))
            trial.set_user_attr("train_max", np.max(train_err))
            trial.set_user_attr("testing_error", np.mean(train_err))
            trial.set_user_attr("test_median", np.median(train_err))
            trial.set_user_attr("test_25_percentile", np.percentile(train_err, 25))
            trial.set_user_attr("test_75_percentile", np.percentile(train_err, 75))
            trial.set_user_attr("test_max", np.max(train_err))

            median_value = np.median(train_err)
            median_absolute_deviation = np.median([abs(number - median_value) for number in train_err])
            trial.set_user_attr("mad", median_absolute_deviation)

            threshold = median_value + 2 * median_absolute_deviation
            trial.set_user_attr("threshold", threshold)

            trial.set_user_attr("test_minus_train_error", np.mean(y_pred)-np.mean(train_err))

            anomaly_count = 0
            for err in y_pred:
                if err > threshold:
                    anomaly_count += 1

            trial.set_user_attr("anomaly_count", anomaly_count)
            trial.set_user_attr("train_packets", training_cutoff)
            trial.set_user_attr("test_packets", len(y_pred))

            FPR = anomaly_count / len(y_pred)
            return FPR

        # Dashboard logic
        search_space = {
            'numAE': [5, 10, 15, 25, 50, 75, 150],
            'learning_rate': [0.0001, 0.0005, 0.001, 0.005, 0.01, 0.05, 0.1],
            'hidden_ratio': [0.25, 0.5, 0.75],
            'FMgrace': [math.floor(0.05*training_cutoff), math.floor(0.10*training_cutoff), math.floor(0.20 * training_cutoff)]
        }
        search_space = {
            'numAE': [50],
            'learning_rate': [0.0001],
            'hidden_ratio': [0.25],
            'FMgrace': [math.floor(0.05*training_cutoff)]
        }
        name = "mad2_hyperopt" + str(training_cutoff) + "learnlowerhigheragainappend"
        study = optuna.create_study(sampler=optuna.samplers.GridSampler(search_space), storage="sqlite:///hyperopt.db", study_name=name)
        study.optimize(objective, n_trials=27)

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write header row
        header = ["Trial Number", "numAE", "learning_rate", "hidden_ratio"]
        ws.append(header)

        # Write trial information
        best_value = float("inf")
        best_row_idx = None  # Track the index of the best row
        for idx, trial in enumerate(study.trials, start=2):  # Start from row 2 to leave room for the header
            trial_params = trial.params
            trial_row = [trial.number, trial_params["numAE"], trial_params["learning_rate"],
                         trial_params["hidden_ratio"], trial.value]
            ws.append(trial_row)

            if trial.value < best_value:
                best_value = trial.value
                best_row_idx = idx

        # Set fill color for the best value row
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        if best_row_idx is not None:
            for cell in ws[best_row_idx]:
                cell.fill = green_fill

        # Save the workbook to a file
        excel_file_path = "output_data/hyperparameter_optimization_results_" + datetime.now().strftime(
            '%d-%m-%Y_%H-%M') + ".xlsx"
        wb.save(excel_file_path)

        print("Results exported to", excel_file_path)
        return study.best_trial

    # Runs a hyperparameter optimization on the supplied dataset, constrained by number of runs and packet limit
    # This version uses KitNET directly instead of running Kitsune as a whole
    def hyper_opt_KitNET_mean(self, feature_path, training_cutoff, total_cutoff):
        def objective(trial):
            numAE = trial.suggest_int('numAE', 0, 200)
            learning_rate = trial.suggest_float('learning_rate', 0, 0.5)
            hidden_ratio = trial.suggest_float('hidden_ratio', 0, 1)
            FMgrace = trial.suggest_int('FMgrace', 0, 500000)

            kit = KitNET(100, max_autoencoder_size=numAE, FM_grace_period=FMgrace, AD_grace_period=math.floor(training_cutoff*0.9), learning_rate=learning_rate, hidden_ratio=hidden_ratio)
            # Load the feature list beforehand to save time
            iter = 0
            with open(feature_path) as fp:
                rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')

                train_err = []
                y_pred = []
                for packet in rd_ft:
                    if packet:
                        packet = packet[0].split(',')
                        packet = [float(element) for element in packet]
                        packet = np.array(packet)
                        if iter % 10000 == 0:
                            print(iter)
                        if iter < total_cutoff:
                            if iter <= training_cutoff:
                                train_err.append(kit.train(packet))
                            else:
                                score = kit.execute(packet)
                                y_pred.append(score)
                            iter += 1
                        else:
                            break
                fp.close()

            trial.set_user_attr("training_error", np.mean(train_err))
            error = np.mean(y_pred)
            print('error')
            print(error)
            return error

        # Dashboard logic
        search_space = {
            'numAE': [5, 10, 15, 25, 50, 75, 150],
            'learning_rate': [0.001, 0.005, 0.01, 0.05, 0.1, 0.13, 0.2],
            'hidden_ratio': [0.25, 0.5, 0.75],
            'FMgrace': [math.floor(0.05*training_cutoff), math.floor(0.10*training_cutoff), math.floor(0.20 * training_cutoff)]
        }
        name = "hyperopt_mean_" + str(total_cutoff)
        study = optuna.create_study(sampler=optuna.samplers.GridSampler(search_space), storage="sqlite:///hyperopt.db", study_name=name)
        study.optimize(objective, n_trials=7*7*3*4)

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write header row
        header = ["Trial Number", "numAE", "learning_rate", "hidden_ratio"]
        ws.append(header)

        # Write trial information
        best_value = float("inf")
        best_row_idx = None  # Track the index of the best row
        for idx, trial in enumerate(study.trials, start=2):  # Start from row 2 to leave room for the header
            trial_params = trial.params
            trial_row = [trial.number, trial_params["numAE"], trial_params["learning_rate"],
                         trial_params["hidden_ratio"], trial.value]
            ws.append(trial_row)

            if trial.value < best_value:
                best_value = trial.value
                best_row_idx = idx

        # Set fill color for the best value row
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        if best_row_idx is not None:
            for cell in ws[best_row_idx]:
                cell.fill = green_fill

        # Save the workbook to a file
        excel_file_path = "output_data/hyperparameter_optimization_results_" + datetime.now().strftime(
            '%d-%m-%Y_%H-%M') + ".xlsx"
        wb.save(excel_file_path)

        print("Results exported to", excel_file_path)
        return study.best_trial

    # Calculates KitNET's SHAP-values for the specified indexes
    def shap_values_builder_separate_train_test_csv(self, train_path, test_path, training_cutoff, test_cutoff, numAE, learning_rate, hidden_ratio):
        self.KitTest = KitNET(100, numAE, math.floor(training_cutoff * 0.1), math.floor(training_cutoff*0.9), learning_rate, hidden_ratio)
        # Load CSV file since it probably will not be too big
        with open(train_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            train_features = []
            for packet in rd_ft:
                if packet:
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    train_features.append(packet)
            fp.close()
        print('Done building train feature array')

        # Load CSV file since it probably will not be too big
        with open(test_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            test_features = []
            iter = 0
            for packet in rd_ft:
                if packet:
                    if iter >= test_cutoff:
                        break
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    test_features.append(packet)
                    iter += 1
            fp.close()
        print('Done building test feature array')

        trainfeaturesNP = np.array(train_features)

        print('Training KitNET')
        self.KitTest.process_batch(trainfeaturesNP[:training_cutoff])
        print("Building SHAP explainer")
        self.explainer = shap.Explainer(self.kitnet_model, trainfeaturesNP[:training_cutoff])
        print("Calculating SHAP values")
        newfeatures = random.sample(test_features, 40)
        # Get 40 random packets from test set
        self.shap_values = self.explainer.shap_values(np.array(newfeatures))
        self.metadata = {
            "filename": train_path,
            "packet_limit": training_cutoff,
            "num_autenc": numAE,
            "FMgrace": math.floor(training_cutoff * 0.1),
            "ADgrace": math.floor(training_cutoff * 0.9),
            "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        }
        return self.shap_values

    def shap_values_builder_from_csv(self, path, training_cutoff, total_cutoff, numAE, learning_rate, hidden_ratio):
        self.KitTest = KitNET(100, numAE, math.floor(training_cutoff * 0.1), math.floor(training_cutoff * 0.9),
                              learning_rate, hidden_ratio)
        # Load CSV file since it probably will not be too big
        with open(path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            features = []
            for packet in rd_ft:
                if packet:
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    features.append(packet)
            fp.close()
        print('Done building feature array')

        featuresNP = np.array(features)

        print('Training KitNET')
        self.KitTest.process_batch(featuresNP[:training_cutoff])
        print("Building SHAP explainer")
        self.explainer = shap.Explainer(self.kitnet_model, featuresNP[:training_cutoff])
        print("Calculating SHAP values")
        newfeatures = features[training_cutoff:total_cutoff]
        newfeatures = random.sample(newfeatures, 100)
        # Get 40 random packets from test set
        self.shap_values = self.explainer.shap_values(np.array(newfeatures))
        self.metadata = {
            "filename": path,
            "packet_limit": total_cutoff,
            "num_autenc": numAE,
            "FMgrace": math.floor(training_cutoff * 0.1),
            "ADgrace": math.floor(training_cutoff * 0.9),
            "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        }
        return self.shap_values

    def run_kitsune_from_feature_csv(self, feature_path, training_cutoff, total_cutoff, numAE, learning_rate, hidden_ratio):
        kit = KitNET(100, numAE, math.floor(training_cutoff * 0.05), training_cutoff * 0.9, learning_rate,
                     hidden_ratio)
        # Load the feature list beforehand to save time
        y_pred = []
        counter = 0
        with open(feature_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')

            for packet in rd_ft:
                if packet:
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    if counter % 10000 == 0:
                        print(counter)
                    if counter < total_cutoff:
                        if counter <= training_cutoff:
                            kit.train(packet)
                        else:
                            score = kit.execute(packet)
                            y_pred.append(score)
                        counter += 1
                    else:
                        break
            fp.close()
            print("Writing anomaly detector to file")
            path = 'pickles/anomDetectorFullDataset.pkl'
            with open(path, 'wb') as f:
                pickle.dump(kit, f)
        return y_pred

    def run_trained_kitsune_from_feature_csv(self, test_path, test_start, test_limit):
        #kit = KitNET(100, 10, math.floor(12000000 * 0.05), math.floor(12000000 * 0.9), 0.30, 0.25)
        # kit = KitNET(100, 50, math.floor(10000000 * 0.05), 10000000, 0.0005, 0.25)
        #
        #
        #
        # counter = 0
        # with open('input_data/features.csv') as fp:
        #     rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
        #
        #     for packet in rd_ft:
        #        if packet:
        #            packet = packet[0].split(',')
        #            packet = [float(element) for element in packet]
        #            packet = np.array(packet)
        #            if counter % 10000 == 0:
        #                print(counter)
        #            if counter < math.floor(700000):
        #                kit.train(packet)
        #                counter += 1
        #            else:
        #                break
        #     fp.close()
        # path = 'pickles/anomDetectorFullDataset2.pkl'
        # with open(path, 'wb') as f:
        #     pickle.dump(kit, f)
        # print('testing')
        # quit()
        # # Load the feature list beforehand to save time
        # counter = 0
        # print(test_start)
        # print(test_limit)
        # with open(test_path) as fp:
        #     rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
        #     resultList = []
        #     for packet in rd_ft:
        #         if counter % 10000 == 0:
        #             print(counter)
        #         if packet and counter > test_start:
        #             print('testing1')
        #             print(counter)
        #             packet = packet[0].split(',')
        #             packet = [float(element) for element in packet]
        #             packet = np.array(packet)
        #             if counter < test_limit:
        #                 print('test2')
        #                 resultList.append(kit.execute(packet))
        #             else:
        #                 break
        #         counter += 1
        #     fp.close()
        #     print("Writing anomaly detector to file")
        #     path = 'pickles/anomDetector.pkl'
        #     with open(path, 'wb') as f:
        #         pickle.dump(kit, f)
        with open("pickles/anomDetectorFullDataset.pkl", 'rb') as f:
            kit = pickle.load(f)

        counter = 0
        results = []
        with open(test_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')

            for packet in rd_ft:
                if packet:
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    if counter % 10000 == 0:
                        print('running: ')
                        print(counter)
                    if counter < test_limit:
                        results.append(kit.execute(packet))
                        counter += 1
                    else:
                        break
            fp.close()
        return results

    def run_trained_kitsune_from_tsv(self, test_path, test_limit):
        path = 'pickles/anomDetectorFullDataset.pkl'
        with open(path, 'rb') as f:
            kit = pickle.load(f)

        # Load the feature list beforehand to save time
        #iter = 0
        #with open(feature_path) as fp:
            #rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')

            #for packet in rd_ft:
            #    if packet:
            #        packet = packet[0].split(',')
            #        packet = [float(element) for element in packet]
            #        packet = np.array(packet)
            #        if iter % 10000 == 0:
            #            print(iter)
            #        if iter < training_cutoff:
            #            kit.train(packet)
            #            iter += 1
            #        else:
            #            break
            #fp.close()
            #print("Writing anomaly detector to file")
            #path = 'pickles/anomDetector.pkl'
            #with open(path, 'wb') as f:
            #    pickle.dump(kit, f)
            #with open(path, 'rb') as f:
            #    newKit = pickle.load(f)

        counter = 0
        results = []
        with open(test_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            for packet in rd_ft:
                if packet and counter > 0:
                    print(counter)
                    features = self.get_features_for_packet(packet)
                    results.append(kit.execute(features))
                counter += 1
            fp.close()
        return results

    def map_results_to_conversation(self, results, pcap_path):
        counter = 0
        conv_dict = {}
        with open(pcap_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            for packet in rd_ft:
                if counter < len(results):
                    if packet:
                        packet = packet[0].split(',')
                        result = results[counter]
                        conv_number = packet[23]
                        if conv_number not in conv_dict:
                            conv_dict[conv_number] = []
                        conv_dict[conv_number].append(result)
                        counter += 1
                else:
                    break
            fp.close()
        return conv_dict

    def run_kitsune_from_feature_pickle(self, feature_path, training_cutoff, total_cutoff, numAE, learning_rate, hidden_ratio, pickle_path=None):
        kit = KitNET(100, numAE, math.floor(training_cutoff * 0.1), math.floor(training_cutoff * 0.9), learning_rate, hidden_ratio)

        #path = 'pickles/anomDetector.pkl'
        #if pickle_path != None:
        #    path = pickle_path
        #with open(path, 'rb') as f:
        #    kit = pickle.load(f)

        # Load the feature list beforehand to save time
        iter = 0
        with open(feature_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')

            y_pred = []
            for packet in rd_ft:
                if packet:
                    packet = packet[0].split(',')
                    packet = [float(element) for element in packet]
                    packet = np.array(packet)
                    if iter % 10000 == 0:
                        print(iter)
                    if iter < total_cutoff:
                        if iter <= training_cutoff:
                            kit.train(packet)
                        else:
                            score = kit.execute(packet)
                            y_pred.append(score)
                        iter += 1
                    else:
                        break
            fp.close()
            print("Writing anomaly detector to file")
            path = 'pickles/anomDetector.pkl'
            with open(path, 'wb') as f:
                pickle.dump(kit, f)
        return y_pred

    def get_features_for_packet(self, packet):
        row = packet
        #row = row[0].strip('][').split(',')
        IPtype = np.nan
        timestamp = row[0]
        framelen = row[1]
        srcIP = ''
        dstIP = ''
        tcpFlags = row[19]
        payload = ''
        # payload = int(row[20])+int(row[21])
        if row[4] != '':  # IPv4
            srcIP = row[4]
            dstIP = row[5]
            IPtype = 0
        elif row[17] != '':  # ipv6
            srcIP = row[17]
            dstIP = row[18]
            IPtype = 1
        srcproto = row[6] + row[
            8]  # UDP or TCP port: the concatenation of the two port strings will will results in an OR "[tcp|udp]"
        dstproto = row[7] + row[9]  # UDP or TCP port
        srcMAC = row[2]
        dstMAC = row[3]
        if srcproto == '':  # it's a L2/L1 level protocol
            if row[12] != '':  # is ARP
                srcproto = 'arp'
                dstproto = 'arp'
                srcIP = row[14]  # src IP (ARP)
                dstIP = row[16]  # dst IP (ARP)
                IPtype = 0
            elif row[10] != '':  # is ICMP
                srcproto = 'icmp'
                dstproto = 'icmp'
                IPtype = 0
            elif srcIP + srcproto + dstIP + dstproto == '':  # some other protocol
                srcIP = row[2]  # src MAC
                dstIP = row[3]  # dst MAC
                ### Extract Features
        try:
            return self.nstat.updateGetStats(IPtype, srcMAC, dstMAC, srcIP, srcproto, dstIP, dstproto,
                                             int(framelen),
                                             float(timestamp), tcpFlags, payload)
        except Exception as e:
            print(e)
            return []

    def most_significant_packets_sampler(self, day):
        root_folder = "."
        attack_types_folder = os.path.join(root_folder, "input_data/attack_types")
        pickles_folder = os.path.join(root_folder, "pickles/output_pickles_packet_basis")

        for attack_type in os.listdir(attack_types_folder):
            if attack_type == f"{day}_features.csv" or attack_type == f"{day}_BENIGN.csv" or not (attack_type.startswith(day) and attack_type.endswith(".csv")):
                continue
            attack_type = attack_type.replace(".csv", "")
            attack_type = attack_type.replace(f"{day}_features_", "")

            # Construct the file names for features and pickle file
            feature_file_name = f"{day}_features_{attack_type}.csv"
            pickle_file_name = f"{day.title()}_{attack_type}_results.pkl"
            feature_file_path = os.path.join(attack_types_folder, feature_file_name)
            pickle_file_path = os.path.join(pickles_folder, pickle_file_name)

            # Check if the pickle file exists
            if not os.path.exists(pickle_file_path):
                continue

            # Load the pickle file containing reconstruction errors
            with open(pickle_file_path, 'rb') as pickle_file:
                reconstruction_errors = pickle.load(pickle_file)

            # Load the corresponding feature CSV file
            features_df = pd.read_csv(feature_file_path, header=None)
            # Sort the errors and get the indices of the 40 highest
            sorted_indices = list(filter(lambda x: x < len(features_df), np.argsort(reconstruction_errors)[-40:]))

            # Extract the significant features
            significant_features = features_df.iloc[sorted_indices]
            # Define the output file name
            output_file_name = f"{day}_features_{attack_type}_most_significant.csv"
            output_file_path = os.path.join(attack_types_folder, output_file_name)
            # Save the significant features to a new CSV file
            significant_features.to_csv(output_file_path, index=False, header=False)

    def shap_values_builder_from_features(self, test_feature_path, benign_feature_path):
        path = 'pickles/anomDetectorFullDataset.pkl'
        with open(path, 'rb') as f:
            kit = pickle.load(f)

        def callKit(featureList):
            results = []
            for features in featureList:
                results.append(kit.execute(features))
            return np.array(results)

        # Load CSV file since it probably will not be too big
        with open(test_feature_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            test_features = []
            for feature in rd_ft:
                feature = feature[0].split(',')
                feature = [float(element) for element in feature]
                feature = np.array(feature)
                test_features.append(feature)
            fp.close()
        print('Done building training feature array')

        # Load CSV file since it probably will not be too big
        with open(benign_feature_path) as fp:
            rd_ft = csv.reader(fp, delimiter="\t", quotechar='"')
            benign_features = []
            for feature in rd_ft:
                feature = feature[0].split(',')
                feature = [float(element) for element in feature]
                feature = np.array(feature)
                benign_features.append(feature)
            fp.close()

        print("Building SHAP explainer")
        explainer = shap.KernelExplainer(callKit, np.array(benign_features[:40]))
        print("Calculating SHAP values")
        self.shap_values = explainer.shap_values(np.array(test_features[:5]))
        return self.shap_values

    # Calculates SHAP-values for each available attack type in a day of the week; writes results to Excel and pickles results
    def shap_documenter(self, day):
        root_folder = "."
        attack_types_folder = os.path.join(root_folder, "input_data/attack_types")
        self.workbook = openpyxl.load_workbook(f'input_data/template_statistics_file.xlsx')

        for attack_type in os.listdir(attack_types_folder):
            if not (attack_type.startswith(day) and attack_type.endswith("most_significant.csv")):
                continue
            attack_type = attack_type.replace(".csv", "")
            attack_type = attack_type.replace(f"{day}_features_", "")
            # Loop over the different Kitsune configs we are going to make
            shap_values = self.shap_values_builder_from_features(
                f"input_data/attack_types/thursday_features_{attack_type}.csv",
                "input_data/sampled_features_100.csv")

            path = f'pickles/output_pickles/{day.title()}_{attack_type}shap_results.pkl'
            with open(path, 'wb') as f:
                pickle.dump(shap_values, f)
            # Could do this with a Regular Expression but I'm a sane person
            self.create_sheet(day, attack_type.replace("most_significant", "").replace("-", "").replace("_", "").replace(" ", ""))
        excel_file = f"output_data/shap_{day}_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
        self.workbook.save(excel_file)
