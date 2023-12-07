import csv
import math
import pickle
from math import floor

import numpy as np
from scapy.utils import rdpcap
from random import sample
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt2

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from KitPlugin import KitPlugin

inputs = {
    "mirai_malicious": {
        "input_path": "input_data/mirai.pcap",
        "input_path_test": "input_data/mirai.pcap",
        "packet_limit": 200000,
        "maxAE": 10,
        "FMgrace": 5000,
        "ADgrace": 50000,
        "training_min": 0,
        "training_max": 60000,
        "testing_min": 140330,
        "testing_max": 140355
    },
    "mirai_benign": {
        "input_path": "input_data/mirai.pcap",
        "input_path_test": "input_data/mirai.pcap",
        "packet_limit": 200000,
        "maxAE": 10,
        "FMgrace": 5000,
        "ADgrace": 50000,
        "training_min": 0,
        "training_max": 60000,
        "testing_min": 70000,
        "testing_max": 70025
    }
}


# KitPlugin = KitPlugin()
# KitPlugin.hyper_opt("input_data/Monday-WorkingHours_10_percent_random.pcap", 100, 1000000)

# Run series of statistics
# KitPlugin.run_series_stats(inputs)
# Get feature list from pickle file
# KitPlugin.feature_loader()
# Train Kitsune on the training data
# KitPlugin.kit_trainer(training_min, training_max)
# Calculate SHAP-values
# KitPlugin.shap_values_builder(training_min, training_max, testing_min, testing_max)
# Pickle SHAP-values
# KitPlugin.shap_values_pickle()
# Load SHAP-values
# KitPlugin.shap_values_loader()
# Calculate shap summary statistics
# KitPlugin.shap_stats_summary_builder(testing_min, testing_max)
# KitPlugin.shap_stats_excel_export()

# Calculate EER and AUC values
# KitPlugin = KitPlugin(input_path="input_data/mirai.pcap", packet_limit=200000, num_autenc=10, FMgrace=5000, ADgrace=50000, learning_rate=0.1, hidden_ratio=0.75)
# KitPlugin.feature_loader()
# KitPlugin.kit_trainer(0, 60000)
# KitPlugin.model_pickle()
# ONLY run this on a mixed batch of benign/malicious samples
# RMSEs = KitPlugin.kit_runner(120000, 122000, normalize=True)
# Labels
# Create an array of zeros with 1622 entries
# zeros = np.zeros(1622)
# Create an array of ones with 378 entries
# ones = np.ones(378)
# Concatenate the two arrays to get the final array
# labels = np.concatenate((zeros, ones))
# KitPlugin.calc_auc_eer(RMSEs, labels)

# KitPlugin = KitPlugin()
# Random sample of 500000 packets, ordered by timestamp
# KitPlugin.random_sample_pcap("input_data/Monday-WorkingHours.pcap", "input_data/Monday-WorkingHours_500k.pcap", 500000)
# KitPlugin.random_sample_pcap("input_data/Monday-WorkingHours.pcap", "input_data/Monday-WorkingHours_1M.pcap", 1000000)

# Out of every 1000 packets, it only keeps the first 100; so, only 10% of packets is kept. It will then do the same for the next 1000 packets
# KitPlugin.interval_sample_pcap("input_data/Monday-WorkingHours.pcap", "input_data/Monday-WorkingHours_10_percent.pcap", 10)

# Sample 10 percent of conversations
# SampleKitPlugin = KitPlugin()
##conversations = SampleKitPlugin.sample_percentage_conversations(10, "input_data/Monday_Split/17_01-18_01.pcapng", "input_data/Monday_Split/17_01-18_01-sample-10.pcap")
# conversations = SampleKitPlugin.conversations_loader('pickles/17_01-18_01_sample_10_conv')
# packets = rdpcap('input_data/Monday_Split/17_01-18_01-sample-10.pcap')
# features = SampleKitPlugin.load_pcap_to_features('input_data/Monday_Split/17_01-18_01-sample-10.pcap')

# print('conversations: '+str(len(conversations)))
# print('packets: '+str(len(packets)))
# print('feature lists: '+str(len(features)))

# del SampleKitPlugin

# NewKitPlugin = KitPlugin('input_data/Monday_Split/17_01-18_01-sample-10.pcap', num_autenc=6, FMgrace=int(0.05*len(features)), int(0.95*len(ADgrace)))

def kitTester(day, attack_type):
    from KitPlugin import KitPlugin
    kitplugin = KitPlugin()
    # print('reading labels file')
    # labels = kitplugin.read_label_file(f'input_data/attack_types/{day}_{attack_type}.csv')
    # iter = 0
    # for label in labels:
    #     if iter == 0:
    #         iter += 1
    #         continue
    #     label.append(str(labels.index(label)))
    # print('sampling packets by conversation')
    # kitplugin.sample_packets_by_conversation(f'input_data/{day.title()}-WorkingHours.pcap.tsv',
    #                                          f'input_data/attack_types/{day}_{attack_type}.pcap.tsv', labels)
    #
    # # Map samples to features of an existing featureList
    # kitplugin.map_packets_to_features(f'input_data/attack_types/{day}_{attack_type}.pcap.tsv',
    #                                   f'input_data/attack_types/{day}_features.csv',
    #                                   f'input_data/attack_types/{day}_features_{attack_type}.csv')
    results = kitplugin.run_trained_kitsune_from_feature_csv(
        f"input_data/attack_types/{day}_features_{attack_type}.csv", 0, np.Inf)
    #plt.title(f'{day.title()}_{attack_type}')
    #plt.plot(results)
    #plt.show()
    with open(f'pickles/output_pickles_packet_basis/{day.title()}_{attack_type}_results.pkl', 'wb') as f:
        pickle.dump(results, f)

    convs = kitplugin.map_results_to_conversation(results, f"input_data/attack_types/{day}_{attack_type}.pcap.tsv")
    print(f"attack: {attack_type}, convs: {len(convs)}")
    maxConvs = []
    for conv in convs:
        maxConvs.append(np.max(convs[conv]))

    path = f'pickles/output_pickles_conv_basis/{day.title()}_{attack_type}_maxConvs.pkl'
    with open(path, 'wb') as f:
        pickle.dump(maxConvs, f)

    return maxConvs

#kitplugin = KitPlugin(input_path="input_data/Monday-WorkingHours.pcap.tsv", packet_limit=np.Inf, num_autenc=50, FMgrace=None, ADgrace=None, learning_rate=0.1, hidden_ratio=0.75)
#kitplugin.feature_builder("input_data/attack_types/monday_features.csv")

def plot_attack_boxplots(data_for_attack_types, include_outliers=True, log_scale=False):
    attack_types = list(data_for_attack_types.keys())
    data = list(data_for_attack_types.values())

    # Storing information about actions taken
    actions_taken = []

    if log_scale:
        data = [[val if val <= 0 else val for val in sublist] for sublist in data]
        actions_taken.append('Log Scale')

    plt.figure(figsize=(10, 6))

    # Set showfliers parameter based on include_outliers argument
    show_outliers = 'outliers' if include_outliers else False
    if include_outliers:
        actions_taken.append("outliers included")
    else:
        actions_taken.append("no outliers included")

    # Create boxplots
    bp = plt.boxplot(data, showfliers=show_outliers)

    # Add sample count annotations to each boxplot
    for i, box in enumerate(bp['boxes']):
        # Calculate the number of samples for each boxplot
        sample_count = len(data[i])

        # Position the text annotation slightly above the boxplot
        conv_count = 0
        if len(data[i]) > 0:
            conv_count = max(data[i])
        plt.text(i + 1, conv_count, f'{sample_count} samples', ha='center', va='bottom')

    plt.xticks(range(1, len(attack_types) + 1), attack_types, rotation=45)
    plt.xlabel('Attack Types')
    plt.ylabel('Values')

    # Include information about actions taken in the title
    title = 'Boxplots for Attack Types'
    if actions_taken:
        title += f' ({", ".join(actions_taken)})'

    plt.title(title)

    # Apply log scale if log_scale argument is True
    if log_scale:
        plt.yscale('log')

    plt.tight_layout()
    plt.figure(figsize=(14, 10))
    plt.show()

def create_attack_barchart_excel(data_for_attack_types):
    attack_types = list(data_for_attack_types.keys())
    data = list(data_for_attack_types.values())

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attack Bar Chart"

    # Write attack types to the first column
    for row, attack_type in enumerate(attack_types, start=1):
        ws.cell(row=row, column=1, value=attack_type)

    # Write data to worksheet
    for col, attack_data in enumerate(data, start=2):
        for row, value in enumerate(attack_data, start=1):
            ws.cell(row=row, column=col, value=value)

    # Create a bar chart
    chart = BarChart()

    # Set chart title and axis labels
    chart.title = "Bar Chart for Attack Types"
    chart.x_axis.title = 'Attack Types'
    chart.y_axis.title = 'Values'

    # Set data for the chart
    values = Reference(ws, min_col=2, min_row=1, max_col=len(attack_types) + 1, max_row=len(attack_types[0]))
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(attack_types) + 1)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # Set the chart position
    ws.add_chart(chart, "E5")

    # Save the workbook
    wb.save("attack_barchart.xlsx")

# attacks1 = ["Infiltration - Portscan"]
# convs = []
# for attack in attacks1:
#     print(attack)
#     convs.append(kitTester("thursday", attack))
#
# attacks2 = ["benign - small", "FTP-Patator", "FTP-Patator - Attempted", "SSH-Patator", "SSH-Patator - Attempted"]
# for attack in attacks2:
#     print(attack)
#     convs.append(kitTester("tuesday", attack))


# attacks = attacks1 + attacks2
# attack_dict = {attack: conv for attack, conv in zip(attacks, convs)}
# plot_attack_boxplots(attack_dict, include_outliers=False, log_scale=False)
# plot_attack_boxplots(attack_dict, include_outliers=False, log_scale=True)
# plot_attack_boxplots(attack_dict, include_outliers=True, log_scale=False)
# plot_attack_boxplots(attack_dict, include_outliers=True, log_scale=True)
# create_attack_barchart_excel(attack_dict)
#kitplugin = KitPlugin(input_path="input_data/Monday-WorkingHours.pcap", packet_limit=np.Inf, num_autenc=50, FMgrace=None, ADgrace=None, learning_rate=0.1, hidden_ratio=0.75)
#kitplugin.feature_builder("input_data/attack_types/monday_features_test.csv", True)

kitplugin = KitPlugin()
#kitplugin.most_significant_packets_sampler("thursday", 0.2667368034640465)
results = kitplugin.shap_documenter("thursday")
# kitplugin.most_significant_packets_sampler("tuesday", 0.2667368034640465)
#results = kitplugin.shap_documenter("wednesday")

#kitplugin.hyper_opt_KitNET("monday_features.csv")
